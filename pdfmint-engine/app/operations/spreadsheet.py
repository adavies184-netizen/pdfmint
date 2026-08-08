from __future__ import annotations

import logging
import math
import os
import re
import shutil
import subprocess
from collections import defaultdict
from pathlib import Path

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ..settings import JOB_TIMEOUT_SECONDS

logger = logging.getLogger("pdfmint.spreadsheet")


def _normalise_number(value: str):
    text = value.strip()
    if not text:
        return ""

    cleaned = text.replace(",", "")
    # Preserve identifiers with leading zeros.
    if re.fullmatch(r"-?\d+", cleaned):
        if len(cleaned.lstrip("-")) > 1 and cleaned.lstrip("-").startswith("0"):
            return text
        try:
            return int(cleaned)
        except ValueError:
            return text

    if re.fullmatch(r"-?\d+\.\d+", cleaned):
        try:
            return float(cleaned)
        except ValueError:
            return text

    # Simple currency / percentage: retain as text rather than guessing formatting.
    return text


def _group_words_into_rows(words, y_tolerance=3.5):
    """Group PyMuPDF words into visually aligned rows."""
    if not words:
        return []

    words = sorted(words, key=lambda w: (w[1], w[0]))
    rows = []
    current = []
    current_y = None

    for word in words:
        x0, y0, x1, y1, text, *_ = word
        cy = (y0 + y1) / 2
        if current_y is None or abs(cy - current_y) <= y_tolerance:
            current.append((x0, x1, text))
            if current_y is None:
                current_y = cy
            else:
                current_y = (current_y * (len(current) - 1) + cy) / len(current)
        else:
            rows.append(sorted(current, key=lambda item: item[0]))
            current = [(x0, x1, text)]
            current_y = cy

    if current:
        rows.append(sorted(current, key=lambda item: item[0]))

    return rows


def _infer_column_starts(rows, page_width):
    """Infer a small set of column anchors from repeated x positions."""
    starts = []
    bucket = max(18.0, page_width / 45.0)
    counts = defaultdict(int)

    for row in rows:
        for x0, _x1, _text in row:
            key = round(x0 / bucket) * bucket
            counts[key] += 1

    # Keep anchors repeated across rows. Always include leftmost text.
    repeated = sorted(x for x, count in counts.items() if count >= 2)

    if not repeated:
        all_x = [item[0] for row in rows for item in row]
        return [min(all_x)] if all_x else [0.0]

    # Merge anchors that are very close.
    for x in repeated:
        if not starts or abs(x - starts[-1]) > bucket * 0.75:
            starts.append(x)
        else:
            starts[-1] = (starts[-1] + x) / 2

    # Avoid creating absurdly wide spreadsheets from noisy PDFs.
    return starts[:20] or [0.0]


def _row_to_cells(row, anchors):
    if not row:
        return []

    cells = [""] * len(anchors)

    for x0, _x1, text in row:
        nearest = min(range(len(anchors)), key=lambda i: abs(x0 - anchors[i]))
        if cells[nearest]:
            cells[nearest] += " " + text
        else:
            cells[nearest] = text

    # Trim empty tail columns.
    while cells and cells[-1] == "":
        cells.pop()

    return [_normalise_number(value) for value in cells]


def pdf_to_xlsx(pdf_path: Path, output_path: Path) -> Path:
    document = fitz.open(pdf_path)
    workbook = Workbook()
    workbook.remove(workbook.active)

    try:
        for page_index, page in enumerate(document, start=1):
            ws = workbook.create_sheet(title=f"Page {page_index}"[:31])

            words = page.get_text("words")
            rows = _group_words_into_rows(words)

            if not rows:
                ws["A1"] = "No extractable text found on this PDF page."
                continue

            anchors = _infer_column_starts(rows, page.rect.width)

            for row_index, row in enumerate(rows, start=1):
                values = _row_to_cells(row, anchors)
                for col_index, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_index, column=col_index, value=value)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Basic readable sizing only; prioritise speed over visual fidelity.
            for col_index in range(1, min(len(anchors), 20) + 1):
                max_len = 0
                for cell in ws[get_column_letter(col_index)]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col_index)].width = min(
                    max(10, max_len + 2), 45
                )

            ws.freeze_panes = "A1"

        if not workbook.sheetnames:
            ws = workbook.create_sheet("Page 1")
            ws["A1"] = "No extractable text found."

        workbook.save(output_path)

    finally:
        document.close()

    if not output_path.exists() or output_path.stat().st_size == 0:
        raise RuntimeError("XLSX conversion did not produce a valid file.")

    logger.info(
        "XLSX created path=%s size_bytes=%s",
        output_path,
        output_path.stat().st_size,
    )
    return output_path


def xlsx_to_xls(xlsx_path: Path, output_dir: Path) -> Path:
    libreoffice_home = output_dir / "libreoffice-calc-home"
    libreoffice_home.mkdir(parents=True, exist_ok=True)

    output_path = output_dir / f"{xlsx_path.stem}.xls"
    if output_path.exists():
        output_path.unlink()

    candidates = [
        "xls",
        'xls:"MS Excel 97"',
        'xls:"MS Excel 95"',
    ]

    failures = []

    for convert_to in candidates:
        profile = libreoffice_home / re.sub(r"[^a-zA-Z0-9]+", "-", convert_to).strip("-").lower()
        if profile.exists():
            shutil.rmtree(profile, ignore_errors=True)
        profile.mkdir(parents=True, exist_ok=True)

        command = [
            "soffice",
            "--headless",
            "--invisible",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--nofirststartwizard",
            "--norestore",
            f"-env:UserInstallation=file://{profile}",
            "--convert-to",
            convert_to,
            "--outdir",
            str(output_dir),
            str(xlsx_path),
        ]

        logger.info("LibreOffice XLS attempt convert_to=%s", convert_to)

        result = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=JOB_TIMEOUT_SECONDS,
            check=False,
            env={
                **os.environ,
                "HOME": str(profile),
                "SAL_USE_VCLPLUGIN": "svp",
                "JAVA_TOOL_OPTIONS": "-Xms16m -Xmx64m",
            },
        )

        if output_path.exists() and output_path.stat().st_size > 0 and result.returncode == 0:
            logger.info("XLS export succeeded with convert_to=%s", convert_to)
            return output_path

        details = (result.stderr or result.stdout or "No XLS file produced.").strip()
        failures.append(f"{convert_to}: {details[:500]}")

    raise RuntimeError(
        "LibreOffice could not export XLS using the available filters. "
        + " | ".join(failures)
    )


def pdf_to_xls(pdf_path: Path, output_dir: Path, base_name: str) -> Path:
    xlsx_path = output_dir / f"{base_name}.xlsx"
    pdf_to_xlsx(pdf_path, xlsx_path)
    return xlsx_to_xls(xlsx_path, output_dir)
